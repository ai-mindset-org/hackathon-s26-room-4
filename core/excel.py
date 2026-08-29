# -*- coding: utf-8 -*-
"""Excel-дайджест по письменному ТЗ заказчика K4UR (docs/interview-K4UR.md):
формат — Excel, поля — наименование, цена предыдущей закупки, текущая цена,
параметры, поставщик и контакты; существенные изменения выделены.

MOQ и форма оплаты в контракте v2 пока не передаются источниками — колонки
оставлены и помечены, чтобы отсутствие данных было видно, а не скрыто.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FILLS = {"red": PatternFill("solid", start_color="FFFFC7CE"),
         "warn": PatternFill("solid", start_color="FFFFEB9C"),
         "deal": PatternFill("solid", start_color="FFC6EFCE")}

HEAD = ["Наименование / параметры", "Поставщик · контакты", "Источник",
        "Цена пред.", "Цена тек.", "Δ%", "Статус", "Валюта", "Наличие",
        "MOQ*", "Оплата*"]


def _sev(pct, red, warn):
    if pct is None:
        return ""
    if pct >= red:
        return "red"
    if pct >= warn:
        return "warn"
    if pct <= -warn:
        return "deal"
    return ""


def write_xlsx(pairs, path, red=10.0, warn=5.0):
    wb = Workbook()
    ws = wb.active
    ws.title = "Дайджест закупки"
    ws.append(HEAD)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for a, b, _events in pairs:
        if b.get("source_status") == "unreachable":
            ws.append(["— источник недоступен —", "", b.get("source", ""),
                       "", "", "", "недоступен", "", "", "", ""])
            ws[ws.max_row][6].fill = FILLS["warn"]
            continue
        prev_ok = a.get("source_status") != "unreachable"
        for sku, bi in b["items"].items():
            ai = a["items"].get(sku) if prev_ok else None
            pct = None
            if (ai and ai["price_status"] == "listed" == bi["price_status"]
                    and ai["currency"] == bi["currency"]
                    and ai["price"] and bi["price"]):
                pct = round((bi["price"] - ai["price"]) / ai["price"] * 100, 1)
            sev = _sev(pct, red, warn)
            status = {"red": f"рост ≥{red:g}% — красный флаг",
                      "warn": "рост заметный", "deal": "подешевело"}.get(
                sev, "" if pct is not None else
                ("по запросу" if bi["price_status"] != "listed" else "новая/без пары"))
            ws.append([bi["title"], bi["shop"], b.get("source", ""),
                       ai["price"] if ai else "", bi["price"],
                       pct if pct is not None else "",
                       status, bi["currency"],
                       "да" if bi["in_stock"] else "нет",
                       bi.get("moq", ""), bi.get("payment", "")])
            if sev:
                for cell in ws[ws.max_row]:
                    cell.fill = FILLS[sev]

    ws.append([])
    ws.append(["* MOQ и форма оплаты источниками пока не отдаются — "
               "поля зарезервированы по ТЗ заказчика K4UR."])
    widths = [58, 34, 22, 11, 11, 8, 26, 8, 9, 8, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path
