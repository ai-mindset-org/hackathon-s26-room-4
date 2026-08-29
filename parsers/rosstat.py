"""Росстат: еженедельные средние потребительские цены по России.

Единственный найденный РОССИЙСКИЙ источник с историей. Магазины отдают только
«сегодня», и первый снимок сравнивать не с чем; здесь каждая неделя 2026 года
уже лежит колонкой, поэтому динамика настоящая и она про российский рынок,
а не про европейский.

Файл: https://rosstat.gov.ru/storage/mediabank/Nedel_sred_cen.xlsx
Листы — по годам (2022…2026), строки — товары, колонки — недели
(«на 12 января», «на 19 января», …). Цена в рублях за кг или за штуку.

⚠ Сертификат. У rosstat.gov.ru сертификат выдан «Russian Trusted Sub CA» —
российским государственным удостоверяющим центром, которого нет в системном
хранилище доверенных. Обычный запрос обрывается на проверке (`http=000`), и
это легко принять за блокировку по стране — мы так и ошиблись сначала.
Проверку сертификата здесь приходится отключать; это безопасно ровно в той
мере, в какой источник публичный и мы только читаем открытую статистику.
"""

from __future__ import annotations

import re
import ssl
import urllib.request
from datetime import date, datetime
from pathlib import Path

from parsers.snapshot import build_snapshot

URL = "https://rosstat.gov.ru/storage/mediabank/Nedel_sred_cen.xlsx"

MONTHS = {"января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5,
          "июня": 6, "июля": 7, "августа": 8, "сентября": 9, "октября": 10,
          "ноября": 11, "декабря": 12}

# Что относится к отделу «Мясо · Рыба · Курица».
MEAT_KEYWORDS = ("говядин", "свинин", "баранин", "куры", "куриц", "цыплят",
                 "индейк", "рыба", "рыбн", "филе", "сельд")


def download(path: str | Path = "/tmp/rosstat-nedel.xlsx") -> Path:
    context = ssl.create_default_context()
    context.check_hostname = False
    context.verify_mode = ssl.CERT_NONE
    request = urllib.request.Request(URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=90, context=context) as response:
        Path(path).write_bytes(response.read())
    return Path(path)


def _week_date(header: str, year: int) -> datetime | None:
    """«на 24 августа» → дата. Заголовки колонок только так и выглядят."""
    match = re.search(r"на\s+(\d{1,2})\s+([а-яё]+)", str(header or ""), re.I)
    if not match:
        return None
    month = MONTHS.get(match.group(2).lower())
    return datetime(year, month, int(match.group(1))) if month else None


def read_year(path: str | Path, year: int = 2026) -> tuple[list[str], list[tuple]]:
    from openpyxl import load_workbook

    sheet = load_workbook(path, read_only=True, data_only=True)[str(year)]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        (i for i, r in enumerate(rows)
         if r and str(r[0] or "").strip().lower().startswith("наименован")), 3)
    return list(rows[header_index]), rows[header_index + 1:]


def snapshots(path: str | Path, year: int = 2026, weeks: int = 2,
              keywords: tuple[str, ...] = MEAT_KEYWORDS) -> list[dict]:
    """Два последних недельных столбца → два снимка контракта v2."""
    header, body = read_year(path, year)

    dated = [(i, _week_date(h, year)) for i, h in enumerate(header)]
    dated = [(i, d) for i, d in dated if d and d.date() <= date.today()]
    if len(dated) < weeks:
        return []

    result = []
    for index, when in dated[-weeks:]:
        items = []
        for row in body:
            name = str(row[0] or "").strip()
            if not name or not any(k in name.lower() for k in keywords):
                continue
            value = row[index] if index < len(row) else None
            price = float(value) if isinstance(value, (int, float)) else None
            items.append({
                "sku": re.sub(r"\s+", "-", re.sub(r"[^\wа-яё ]", " ",
                                                  name.lower()))[:60].strip("-"),
                "shop": "rosstat.gov.ru",
                "title": name,
                "price": price,
                "currency": "RUB",
                "price_status": "listed" if price is not None else "unknown",
                "item_status": "ok" if price is not None else "not_found",
                "in_stock": price is not None,
            })
        result.append(build_snapshot(
            "эталон Росстат · средняя по РФ (не поставщик)", items,
            taken_at=when.isoformat(timespec="seconds")))
    return result


def main(argv: list[str]) -> int:
    import json
    import sys

    for stream in (sys.stdout, sys.stderr):
        if hasattr(stream, "reconfigure"):
            stream.reconfigure(encoding="utf-8")

    out = Path(argv[1] if len(argv) > 1 else "departments/myaso/data")
    out.mkdir(parents=True, exist_ok=True)

    pair = snapshots(download())
    if len(pair) < 2:
        print("меньше двух недель — сравнивать не с чем", file=sys.stderr)
        return 1

    for snap in pair:
        week = snap["taken_at"][:10]
        (out / f"{week}-rosstat.json").write_text(
            json.dumps(snap, ensure_ascii=False, indent=1), encoding="utf-8")

    before, after = pair
    print(f"Росстат, Россия: {before['taken_at'][:10]} → {after['taken_at'][:10]}")
    for sku, item in after["items"].items():
        was = before["items"].get(sku, {}).get("price")
        now = item["price"]
        change = f'{(now / was - 1) * 100:+6.2f}%' if was and now else "     —"
        print(f"  {item['title'][:44]:<44} {was:>8} → {now:>8} ₽  {change}")
    return 0


if __name__ == "__main__":
    import sys
    raise SystemExit(main(sys.argv))
